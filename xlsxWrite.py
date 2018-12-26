from openpyxl import Workbook
from productNew import product_obj

wb = Workbook()

ws = wb.create_sheet('Product', 0)
ws = wb.active

productkeys = ['Product Name', 'Available Quantity', 'Type', 'Profit/Loss', 'Valuation', 'Purchase Value', 'Sell Value']

for i in range(65, 72):
    ws.column_dimensions[chr(i)].width = 16

ws.append(productkeys)

csvvalues = product_obj.product
ws.append([csvvalues['name'], csvvalues['available_qty'], csvvalues['type'], csvvalues['profit_loss'],
           csvvalues['valuation_price'], csvvalues['purchase_value'], csvvalues['sell_value']])

ws = wb.create_sheet('Purchase', 1)

for i in range(65, 73):
    ws.column_dimensions[chr(i)].width = 16

productkeys = ['Purchase Order', 'Date', 'Vendor Name', 'Quantity', 'Price', 'Total GST', 'Total Price', 'Unit Price']
ws.append(productkeys)

for orders in csvvalues['purchase_order']:
    data = product_obj.purchase_details[orders]
    ws.append(
        [orders, data['date'], data['name'], data['quantity'], data['price'], data['total_tax'], data['total_price'],
         data['unit_price']])

ws = wb.create_sheet('Sell', 2)

for i in range(65, 73):
    ws.column_dimensions[chr(i)].width = 16

productkeys = ['Sell Order', 'Date', 'Customer Name', 'Quantity', 'Price', 'Total GST', 'Total Price', 'Unit Price']
ws.append(productkeys)

for orders in csvvalues['sell_order']:
    data = product_obj.sell_details[orders]
    ws.append(
        [orders, data['date'], data['name'], data['quantity'], data['price'], data['total_tax'], data['total_price'],
         data['unit_price']])

wb.save('/home/emipro/Desktop/Training/product.xlsx')
